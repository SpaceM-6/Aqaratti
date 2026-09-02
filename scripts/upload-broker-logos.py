"""
يرفع شعارات الوسطاء (3,630 صورة) من مجلد logos/ إلى Supabase Storage
(bucket: broker-logos)، ثم يحدّث عمود logo_url في جدول brokers لكل وسيط
له شعار، بمطابقة رقم الترخيص (license) الموجود في
dubai_brokers_WITH_LOGOS_only.xlsx مع الجدول. الوسطاء بدون شعار في هذا
الملف يبقى logo_url عندهم كما هو (NULL) بدون أي تعديل.

قبل التشغيل: .env يجب أن يحتوي SUPABASE_URL و SUPABASE_SERVICE_ROLE_KEY.

التشغيل:
  python scripts/upload-broker-logos.py
"""
import concurrent.futures
import os
import sys
import threading
import urllib.parse

import requests
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LOGOS_DIR = os.path.join(ROOT_DIR, "logos")
EXCEL_FILE = os.path.join(ROOT_DIR, "dubai_brokers_WITH_LOGOS_only.xlsx")
BUCKET = "broker-logos"
MAX_WORKERS = 8

print_lock = threading.Lock()


def load_env_file(path):
    env = {}
    if not os.path.exists(path):
        return env
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def get_config():
    env = load_env_file(os.path.join(ROOT_DIR, ".env"))
    supabase_url = (os.environ.get("SUPABASE_URL") or env.get("SUPABASE_URL", "")).rstrip("/")
    service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        raise SystemExit("يجب ضبط SUPABASE_URL و SUPABASE_SERVICE_ROLE_KEY في ملف .env بجذر المشروع.")
    return supabase_url, service_key


def ensure_bucket(session, url):
    resp = session.post(f"{url}/storage/v1/bucket", json={"id": BUCKET, "name": BUCKET, "public": True})
    if resp.status_code in (200, 201):
        print(f"تم إنشاء bucket '{BUCKET}'.")
    elif resp.status_code == 400 and "already exists" in resp.text.lower():
        print(f"bucket '{BUCKET}' موجود بالفعل.")
    else:
        print(f"ملاحظة عند إنشاء bucket: {resp.status_code} {resp.text[:200]}")


def upload_logo(session, url, filename, filepath):
    with open(filepath, "rb") as f:
        data = f.read()
    resp = session.post(
        f"{url}/storage/v1/object/{BUCKET}/{filename}",
        data=data,
        headers={"Content-Type": "image/jpeg", "x-upsert": "true"},
        timeout=30,
    )
    if resp.status_code >= 300:
        raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:200]}")
    return f"{url}/storage/v1/object/public/{BUCKET}/{filename}"


def update_broker_logo(session, url, license_no, logo_url):
    encoded_license = urllib.parse.quote(license_no)
    resp = session.patch(
        f"{url}/rest/v1/brokers?license=eq.{encoded_license}",
        json={"logo_url": logo_url},
        headers={"Content-Type": "application/json", "Prefer": "return=representation"},
        timeout=30,
    )
    if resp.status_code >= 300:
        return "error", resp.text[:200]
    rows = resp.json()
    return ("matched" if rows else "not_found"), None


def process_row(session, supabase_url, license_no, logo_filename):
    filepath = os.path.join(LOGOS_DIR, logo_filename)
    if not os.path.exists(filepath):
        return "missing_file", logo_filename

    try:
        logo_url = upload_logo(session, supabase_url, logo_filename, filepath)
    except Exception as e:
        return "upload_error", f"{logo_filename}: {e}"

    status, detail = update_broker_logo(session, supabase_url, license_no, logo_url)
    if status == "not_found":
        return "not_found", license_no
    if status == "error":
        return "update_error", f"{license_no}: {detail}"
    return "matched", None


def main():
    supabase_url, service_key = get_config()

    session = requests.Session()
    session.headers.update({"apikey": service_key, "Authorization": f"Bearer {service_key}"})

    ensure_bucket(session, supabase_url)

    wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    idx = {name: i for i, name in enumerate(header)}
    license_idx = idx["license"]
    logo_filename_idx = idx["logo_filename"]

    # نأخذ أول ظهور لكل رقم ترخيص فقط - بعض الفروع تشارك نفس ترخيص المكتب الرئيسي،
    # وجدول brokers فيه صف واحد لكل ترخيص، فمعالجة تكرارات بالتوازي كانت تسبب
    # تعارضاً (Race Condition) يكتب شعار فرع خطأً فوق شعار المكتب الرئيسي.
    seen_licenses = set()
    jobs = []
    for row in rows_iter:
        if row is None:
            continue
        license_no = row[license_idx]
        logo_filename = row[logo_filename_idx]
        if not license_no or not logo_filename:
            continue
        license_no = str(license_no).strip()
        if license_no in seen_licenses:
            continue
        seen_licenses.add(license_no)
        jobs.append((license_no, str(logo_filename).strip()))
    wb.close()

    print(f"عدد السجلات في الملف: {len(jobs)}")

    counters = {"matched": 0, "not_found": 0, "missing_file": 0, "upload_error": 0, "update_error": 0}
    done = 0

    def run(job):
        license_no, logo_filename = job
        return process_row(session, supabase_url, license_no, logo_filename)

    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        for status, detail in executor.map(run, jobs):
            with print_lock:
                counters[status] = counters.get(status, 0) + 1
                done += 1
                if detail and status not in ("matched",):
                    print(f"  [{status}] {detail}")
                if done % 200 == 0:
                    print(f"تقدّم: {done} / {len(jobs)}")

    print("---")
    print(f"تم الرفع والتحديث: {counters.get('matched', 0)}")
    print(f"رقم ترخيص غير موجود في جدول brokers: {counters.get('not_found', 0)}")
    print(f"ملف الصورة غير موجود محلياً: {counters.get('missing_file', 0)}")
    print(f"أخطاء رفع: {counters.get('upload_error', 0)}")
    print(f"أخطاء تحديث: {counters.get('update_error', 0)}")


if __name__ == "__main__":
    main()
